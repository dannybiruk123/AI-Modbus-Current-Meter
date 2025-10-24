import threading
import queue
import time
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from pymodbus.client import ModbusSerialClient

# pyserial: список портов и описания
try:
    from serial.tools import list_ports
except Exception:
    list_ports = None

# openpyxl для Excel
try:
    from openpyxl import Workbook
    from openpyxl.chart import Reference
    from openpyxl.chart import ScatterChart, Series
    from openpyxl.drawing.image import Image as ExcelImage
except Exception:
    Workbook = None  # сообщим пользователю при попытке сохранения

# Pillow для сохранения Canvas -> PNG (опционально)
try:
    from PIL import ImageGrab
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False


# ---------- ПОДКЛЮЧЕНИЕ (как в твоей рабочей версии) ----------
class ModbusInputReader:
    def __init__(self, port="COM3", baudrate=115200, unit_id=2, timeout=0.1):
        self.port = port
        self.baudrate = baudrate
        self.unit_id = unit_id
        self.timeout = timeout
        self.client = None

    def connect(self):
        self.client = ModbusSerialClient(
            port=self.port,
            baudrate=self.baudrate,
            timeout=self.timeout,
            parity='N',
            stopbits=1,
            bytesize=8
        )
        return self.client.connect()

    def disconnect(self):
        if self.client:
            try:
                self.client.close()
            finally:
                self.client = None

    def read_ch(self, address=0, count=1):
        # В pymodbus 3.x проверяем client.connected
        if not self.client or not getattr(self.client, "connected", False):
            return None
        try:
            result = self.client.read_input_registers(address, count, slave=self.unit_id)
            if not result.isError():
                return result.registers
            return None
        except Exception:
            return None

    # запись hold-регистров для управления каналами
    def write_reg(self, address, value):
        if not self.client or not getattr(self.client, "connected", False):
            return False
        try:
            res = self.client.write_register(address, int(value) & 0xFFFF, slave=self.unit_id)
            return not res.isError()
        except Exception:
            return False

class ModbusGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Modbus — пусковые токи: 1 канал, 10 мс, Амперы, Excel")
        self.geometry("940x740")
        self.minsize(900, 700)

        # Межпоточная очередь для обновления UI
        self.data_queue = queue.Queue()
        self.worker_thread = None
        self.stop_event = threading.Event()
        self.reader = None
        self.is_connected = False

        # Частота/окно тренда (по умолчанию 60с; можно менять в UI)
        self.sample_interval_ms = 10
        self.window_seconds = 60
        self.points_per_second = int(1000 / self.sample_interval_ms)  # 100 Гц
        self._set_trend_window(self.window_seconds)

        # Буферы
        self.series_all = []     # [(datetime, A_int)] — все точки (целые А)
        self.trend_buffer = []   # [(datetime, [8 значений в А (int/None)])] — для отрисовки

        # Отображение портов
        self.port_display_to_device = {}
        self.port_display_to_info = {}

        # Поля подключения
        self.port_var = tk.StringVar(value="COM3")
        self.port_display_var = tk.StringVar(value="")
        self.baud_var = tk.StringVar(value="115200")
        self.unit_var = tk.StringVar(value="2")
        self.timeout_var = tk.StringVar(value="0.1")

        # Быстрый канал
        self.fast_channel_var = tk.StringVar(value="1")  # 1..8

        # Управление каналами: HR[49] разрешение записи, HR[8..15] on/off
        self.unlock_reg = 49
        self.ctrl_base = 8      # 1-й канал в HR8, последний в HR15
        self._muted = False     # помним, что отключали остальные каналы

        # анти-«ложный старт», антиспайк
        self._warmup_to_skip = 10   # пропускаем первые N валидных выборок
        self._last_a_int = None     # для анти-спайка

        # таймер замера
        self._start_time = None
        self._timer_job = None

        self._build_ui()
        self._init_styles()
        self.refresh_ports()
        self._schedule_queue_pump()

    # ---------- UI / Стили ----------
    def _init_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use('clam')
        except tk.TclError:
            pass
        style.configure("TButton", padding=6)
        style.configure("Accent.TButton", padding=6)
        style.map("Accent.TButton", background=[('active', '!disabled', '#e0e0e0')])
        style.configure("Header.TLabel", font=("Segoe UI", 11, "bold"))
        style.configure("Value.TLabel", font=("Consolas", 14))
        style.configure("Timer.TLabel", font=("Consolas", 14))
        style.configure("Hint.TLabel", foreground="#555")
        style.configure("Ok.TLabel", foreground="#17803D")
        style.configure("Warn.TLabel", foreground="#C07A00")
        style.configure("Bad.TLabel", foreground="#B00020")

    def _build_ui(self):
        root = ttk.Frame(self, padding=10)
        root.pack(fill=tk.BOTH, expand=True)

        # --- Подключение ---
        top = ttk.LabelFrame(root, text="Подключение", padding=10)
        top.pack(fill=tk.X)

        row1 = ttk.Frame(top); row1.pack(fill=tk.X)
        ttk.Label(row1, text="Порт:").pack(side=tk.LEFT, padx=(0, 6))
        self.port_combo = ttk.Combobox(
            row1, textvariable=self.port_display_var, state="readonly",
            width=52, postcommand=self._sync_ports_if_changed
        )
        self.port_combo.pack(side=tk.LEFT, padx=(0, 6))
        self.port_combo.bind("<<ComboboxSelected>>", self.on_port_selected)
        ttk.Button(row1, text="Обновить", command=self.refresh_ports).pack(side=tk.LEFT)

        self.dev_label_var = tk.StringVar(value="—")
        row2 = ttk.Frame(top); row2.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(row2, text="Устройство:", style="Header.TLabel").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Label(row2, textvariable=self.dev_label_var, style="Hint.TLabel").pack(side=tk.LEFT, fill=tk.X, expand=True)

        row3 = ttk.Frame(top); row3.pack(fill=tk.X, pady=(8, 0))

        def add_labeled(entry_parent, label, var, width=12):
            f = ttk.Frame(entry_parent)
            ttk.Label(f, text=label).pack(side=tk.LEFT, padx=(0, 6))
            ttk.Entry(f, textvariable=var, width=width).pack(side=tk.LEFT)
            return f

        add_labeled(row3, "Скорость:", self.baud_var, 10).pack(side=tk.LEFT, padx=(0, 12))
        add_labeled(row3, "Unit ID:", self.unit_var, 6).pack(side=tk.LEFT, padx=(0, 12))
        add_labeled(row3, "Таймаут (с):", self.timeout_var, 6).pack(side=tk.LEFT, padx=(0, 12))

        ttk.Label(row3, text="Канал:").pack(side=tk.LEFT, padx=(12, 6))
        self.fast_combo = ttk.Combobox(row3, state="readonly", width=4,
                                       values=[str(i) for i in range(1, 9)],
                                       textvariable=self.fast_channel_var)
        self.fast_combo.pack(side=tk.LEFT)

        # окно тренда
        ttk.Label(row3, text="  Окно тренда:").pack(side=tk.LEFT, padx=(16, 6))
        self.trend_window_combo = ttk.Combobox(row3, state="readonly", width=7,
                                               values=["10 s", "30 s", "60 s", "5 min"])
        self.trend_window_combo.set("60 s")
        self.trend_window_combo.bind("<<ComboboxSelected>>", self._on_trend_window_change)
        self.trend_window_combo.pack(side=tk.LEFT)

        # --- Кнопки управления + таймер ---
        btns = ttk.Frame(root)
        btns.pack(fill=tk.X, pady=(10, 0))

        self.connect_btn = ttk.Button(btns, text="Подключиться", command=self.on_connect, style="Accent.TButton")
        self.disconnect_btn = ttk.Button(btns, text="Отключиться", command=self.on_disconnect, state=tk.DISABLED)
        self.start_btn = ttk.Button(btns, text="Старт опроса", command=self.on_start, state=tk.DISABLED)
        self.stop_btn = ttk.Button(btns, text="Стоп опроса", command=self.on_stop, state=tk.DISABLED)
        self.save_btn = ttk.Button(btns, text="Сохранить в Excel", command=self.on_save_excel)
        self.clear_btn = ttk.Button(btns, text="Очистить буфер", command=self.on_clear_data)

        self.connect_btn.pack(side=tk.LEFT)
        self.disconnect_btn.pack(side=tk.LEFT, padx=(8, 0))
        self.start_btn.pack(side=tk.LEFT, padx=(8, 0))
        self.stop_btn.pack(side=tk.LEFT, padx=(8, 0))
        self.save_btn.pack(side=tk.LEFT, padx=(8, 0))
        self.clear_btn.pack(side=tk.LEFT, padx=(8, 0))

        # таймер + статус справа
        right_box = ttk.Frame(btns); right_box.pack(side=tk.RIGHT)
        self.timer_var = tk.StringVar(value="00:00:00")
        ttk.Label(right_box, text="Время замера:", style="Header.TLabel").pack(side=tk.LEFT, padx=(0, 6))
        self.timer_label = ttk.Label(right_box, textvariable=self.timer_var, style="Timer.TLabel")
        self.timer_label.pack(side=tk.LEFT, padx=(0, 12))

        self.status_var = tk.StringVar(value="Отключено")
        ttk.Label(right_box, textvariable=self.status_var).pack(side=tk.LEFT)

        # --- Текущее значение (в А) ---
        mid = ttk.LabelFrame(root, text="Текущее значение (А)", padding=10)
        mid.pack(fill=tk.X, pady=(10, 0))
        rowv = ttk.Frame(mid); rowv.pack(fill=tk.X)
        ttk.Label(rowv, text="Канал:", style="Header.TLabel").pack(side=tk.LEFT, padx=(0, 8))
        self.cur_ch_label = ttk.Label(rowv, textvariable=self.fast_channel_var, style="Header.TLabel")
        self.cur_ch_label.pack(side=tk.LEFT)
        ttk.Label(rowv, text="  Значение:", style="Header.TLabel").pack(side=tk.LEFT, padx=(16, 8))
        self.value_label = ttk.Label(rowv, text="—", style="Value.TLabel")
        self.value_label.pack(side=tk.LEFT)

        # --- Тренд (Canvas) ---
        trend_frame = ttk.LabelFrame(root, text="Тренд (А)", padding=10)
        trend_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        self.trend_canvas = tk.Canvas(trend_frame, height=380, background="#ffffff",
                                      highlightthickness=1, highlightbackground="#cccccc")
        self.trend_canvas.pack(fill=tk.BOTH, expand=True)
        self.trend_canvas.bind("<Configure>", lambda e: self._redraw_trend())

        ttk.Label(trend_frame,
                  text="Сетка: 1 с (основная), 100 мс (минорная). Значения — целые Амперы. Окно — выбирается сверху.",
                  style="Hint.TLabel").pack(side=tk.LEFT, pady=(6, 0))

        # Закрытие окна — корректная остановка/отключение
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    # ---------- helpers ----------
    def _set_trend_window(self, seconds):
        self.window_seconds = int(seconds)
        self.trend_buffer_max = self.window_seconds * self.points_per_second  # rolling, перезаполняем

    def _on_trend_window_change(self, *_):
        label = self.trend_window_combo.get()
        if label == "10 s": self._set_trend_window(10)
        elif label == "30 s": self._set_trend_window(30)
        elif label == "60 s": self._set_trend_window(60)
        elif label == "5 min": self._set_trend_window(300)
        self._redraw_trend()

    # ---------- Порты ----------
    def _format_port_display(self, p):
        try:
            vid = f"{p.vid:04X}" if p.vid is not None else None
            pid = f"{p.pid:04X}" if p.pid is not None else None
        except Exception:
            vid = pid = None
        desc = p.description or "Неизвестное устройство"
        extras = []
        if vid and pid: extras.append(f"VID:PID={vid}:{pid}")
        if getattr(p, "serial_number", None): extras.append(f"SN={p.serial_number}")
        if getattr(p, "manufacturer", None): extras.append(f"Manuf={p.manufacturer}")
        if getattr(p, "product", None): extras.append(f"Prod={p.product}")
        label = f"{p.device} — {desc}"
        if extras: label += " (" + ", ".join(extras) + ")"
        return label

    def _port_long_info(self, p):
        fields = [f"Порт: {p.device}"]
        if p.description: fields.append(f"Описание: {p.description}")
        if p.hwid: fields.append(f"HWID: {p.hwid}")
        if getattr(p, "manufacturer", None): fields.append(f"Производитель: {p.manufacturer}")
        if getattr(p, "product", None): fields.append(f"Продукт: {p.product}")
        if getattr(p, "serial_number", None): fields.append(f"Серийный номер: {p.serial_number}")
        try:
            if p.vid is not None and p.pid is not None:
                fields.append(f"VID:PID={p.vid:04X}:{p.pid:04X}")
        except Exception:
            pass
        return " | ".join(fields) if fields else "—"

    def refresh_ports(self):
        items = []
        self.port_display_to_device.clear()
        self.port_display_to_info.clear()

        if list_ports is None:
            self.port_combo.configure(values=[])
            self.port_display_var.set("")
            self.dev_label_var.set("pyserial не доступен: невозможно получить список портов")
            return

        ports = list(list_ports.comports())
        for p in ports:
            label = self._format_port_display(p)
            items.append(label)
            self.port_display_to_device[label] = p.device
            self.port_display_to_info[label] = self._port_long_info(p)

        if not items:
            self.port_combo.configure(values=["(Порты не найдены)"])
            self.port_display_var.set("(Порты не найдены)")
            self.dev_label_var.set("Подключите устройство и нажмите «Обновить».")
        else:
            self.port_combo.configure(values=items)
            current = self.port_display_var.get()
            if current in items:
                self.on_port_selected()
            else:
                self.port_display_var.set(items[0])
                self.on_port_selected()

    def _sync_ports_if_changed(self):
        self.refresh_ports()

    def on_port_selected(self, *_):
        disp = self.port_display_var.get()
        device = self.port_display_to_device.get(disp)
        info = self.port_display_to_info.get(disp, "—")
        if device:
            self.port_var.set(device)
        self.dev_label_var.set(info)

    # ---------- Очередь UI ----------
    def _schedule_queue_pump(self):
        try:
            while True:
                item = self.data_queue.get_nowait()
                self._process_queue_item(item)
        except queue.Empty:
            pass
        self.after(50, self._schedule_queue_pump)

    def _process_queue_item(self, item):
        kind, payload = item
        if kind == "values":
            ts, a_int = payload
            # текущий лейбл
            self.value_label.configure(text=("—" if a_int is None else f"{a_int:d} A"))
            # буфер тренда: только активный канал в своей позиции
            a_line = [None]*8
            ch = max(1, min(8, int(self.fast_channel_var.get()))) - 1
            a_line[ch] = a_int
            self.trend_buffer.append((ts, a_line))
            if len(self.trend_buffer) > self.trend_buffer_max:
                self.trend_buffer = self.trend_buffer[-self.trend_buffer_max:]
            self._redraw_trend()
        elif kind == "status":
            self.status_var.set(payload)

    # ---------- Конверсия CODE -> А + антиспайк ----------
    @staticmethod
    def _code_to_mA(code):
        if code is None:
            return None
        return 0.000610 * float(code)  # стандарт: 0.000610 мА/код

    @staticmethod
    def _mA_to_A(I_mA):
        if I_mA is None:
            return None
        # линейка: 4..20 мА -> 0..63 А
        return (I_mA - 4.0) * (63.0 / 16.0)

    def _code_to_A_int(self, raw_code):
        if raw_code is None:
            return None
        code = int(raw_code) & 0xFFFF

        # отбрасываем явный мусор
        if code in (0xFFFF, 0xFFFE, 0x0000, 0x7FFF):
            return None

        I = self._code_to_mA(code)
        # детект обрыва/невалидных: ниже 3.8 мА и выше 30 мА — считаем шумом
        if I is None or I < 3.8 or I > 30.0:
            return None

        A = self._mA_to_A(I)
        if A is None:
            return None

        # кламп 0..63 и округление
        A = max(0.0, min(63.0, A))
        A_int = int(round(A))

        # антиспайк: если скачок >10 А к последнему валидному — игнорируем (берём прошлое)
        if self._last_a_int is not None and abs(A_int - self._last_a_int) > 10:
            return self._last_a_int

        self._last_a_int = A_int
        return A_int

    # ---------- Таймер ----------
    def _timer_tick(self):
        if self._start_time is None:
            self.timer_var.set("00:00:00")
            return
        delta = datetime.datetime.now() - self._start_time
        total = int(delta.total_seconds())
        hh = total // 3600
        mm = (total % 3600) // 60
        ss = total % 60
        self.timer_var.set(f"{hh:02d}:{mm:02d}:{ss:02d}")
        self._timer_job = self.after(200, self._timer_tick)

    def _timer_start(self):
        self._start_time = datetime.datetime.now()
        if self._timer_job is not None:
            try: self.after_cancel(self._timer_job)
            except Exception: pass
        self._timer_job = self.after(0, self._timer_tick)

    def _timer_stop(self):
        if self._timer_job is not None:
            try: self.after_cancel(self._timer_job)
            except Exception: pass
        self._timer_job = None
        self._start_time = None
        self.timer_var.set("00:00:00")

    # ---------- Графика ----------
    def _redraw_trend(self):
        c = self.trend_canvas
        c.delete("all")
        w = c.winfo_width()
        h = c.winfo_height()
        if w < 60 or h < 60:
            return

        # рамка
        c.create_rectangle(1, 1, w-2, h-2, outline="#dddddd")

        # горизонтальная сетка
        for i in range(1, 5):
            y = 1 + i*(h-2)/5
            c.create_line(1, y, w-2, y, fill="#eeeeee")

        if len(self.trend_buffer) < 2:
            c.create_text(w/2, h/2, text="Недостаточно данных для тренда", fill="#888888")
            return

        # окно по времени
        data_window = self.trend_buffer[-self.trend_buffer_max:]
        n_full = len(data_window)
        left_pad = 46
        right_pad = 10
        plot_w = max(10, w - left_pad - right_pad)

        # собираем серию активного канала
        series_vals = []
        for _, vals in data_window:
            v = next((x for x in vals if isinstance(x, int)), None)
            series_vals.append(v)

        nums = [v for v in series_vals if isinstance(v, int)]
        if not nums:
            c.create_text(w/2, h/2, text="Нет числовых данных", fill="#888888")
            return
        vmin, vmax = min(nums), max(nums)
        if vmin == vmax:
            vmin -= 1; vmax += 1
        else:
            pad = max(1, int(round((vmax - vmin) * 0.05)))
            vmin -= pad; vmax += pad

        def proj_y(v):
            return 1 + (h-2) * (1 - (v - vmin) / max(1, (vmax - vmin)))

        # вертикальная сетка: 1с + минорные 100 мс
        for sec in range(0, self.window_seconds + 1):
            idx = n_full - 1 - sec * self.points_per_second
            if 0 <= idx < n_full:
                x = left_pad + plot_w * (idx / max(1, n_full - 1))
                c.create_line(x, 1, x, h-2, fill="#f0f0f0")
                if sec % 5 == 0:
                    c.create_text(x, h-12, text=f"-{sec}s", fill="#666666", anchor="e")
            for ms in range(1, 10):
                k = sec * self.points_per_second + ms * (self.points_per_second // 10)
                idx2 = n_full - 1 - k
                if 0 <= idx2 < n_full:
                    x2 = left_pad + plot_w * (idx2 / max(1, n_full - 1))
                    c.create_line(x2, 1, x2, h-2, fill="#fafafa")

        # адаптивная децимация — рисуем не более чем в 1 точку на пиксель
        step = max(1, int((n_full + plot_w - 1) // plot_w))
        pts = []
        for i in range(0, n_full, step):
            v = series_vals[i]
            if isinstance(v, int):
                x = left_pad + plot_w * (i / max(1, n_full - 1))
                y = proj_y(v)
                pts.append((x, y))
            else:
                if len(pts) >= 2:
                    for j in range(1, len(pts)):
                        c.create_line(pts[j-1][0], pts[j-1][1], pts[j][0], pts[j][1], fill="#1f77b4", width=2)
                pts = []
        if len(pts) >= 2:
            for j in range(1, len(pts)):
                c.create_line(pts[j-1][0], pts[j-1][1], pts[j][0], pts[j][1], fill="#1f77b4", width=2)

        c.create_text(6, 12, text=f"max={vmax:d} A", fill="#666666", anchor="w")
        c.create_text(6, h-12, text=f"min={vmin:d} A", fill="#666666", anchor="w")

    # ---------- Подключение (как было) ----------
    def on_connect(self):
        if self.is_connected:
            return

        port = self.port_var.get().strip()
        if not port or "(Порты не найдены)" in self.port_display_var.get():
            messagebox.showerror("Ошибка", "Выберите COM-порт.")
            return
        try:
            baudrate = int(self.baud_var.get().strip())
            unit_id = int(self.unit_var.get().strip())
            timeout = float(self.timeout_var.get().strip())
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте числовые поля подключения.")
            return

        self.reader = ModbusInputReader(port=port, baudrate=baudrate, unit_id=unit_id, timeout=timeout)
        if not self.reader.connect():
            self.status_var.set("Ошибка подключения")
            messagebox.showerror("Ошибка подключения", "Не удалось открыть порт/подключиться.")
            return

        # Рукопожатие: 3 попытки прочитать выбранный канал
        fast_ch = max(1, min(8, int(self.fast_channel_var.get())))
        ok = False
        for _ in range(3):
            vals = self.reader.read_ch(fast_ch - 1, 1)
            if vals is not None and len(vals) >= 1:
                ok = True
                break
            time.sleep(0.05)

        if not ok:
            self.reader.disconnect()
            self.status_var.set("Отключено")
            messagebox.showerror("Ошибка подключения", "Нет ответа от устройства (проверьте Unit ID/проводку).")
            return

        self.is_connected = True
        self.status_var.set(f"Подключено: {port} @ {baudrate}, Unit {unit_id}")
        self.connect_btn.configure(state=tk.DISABLED)
        self.disconnect_btn.configure(state=tk.NORMAL)
        self.start_btn.configure(state=tk.NORMAL)

    def on_disconnect(self):
        # Останавливаем опрос, если идёт
        self.on_stop()
        # восстановим каналы, если мы их трогали
        self._restore_channels()
        if self.reader:
            try:
                self.reader.disconnect()
            except Exception:
                pass
        self.is_connected = False
        self.status_var.set("Отключено")
        self.connect_btn.configure(state=tk.NORMAL)
        self.disconnect_btn.configure(state=tk.DISABLED)
        self.start_btn.configure(state=tk.DISABLED)

    # ---------- Управление каналами ----------
    def _mute_others_and_enable_selected(self, fast_ch):
        """HR49=1; HR8..15=0; HR(7+fast_ch)=1; HR49=0"""
        if not (self.reader and self.is_connected):
            return
        try:
            self.reader.write_reg(self.unlock_reg, 1)
            # выключить всех
            for i in range(8):
                self.reader.write_reg(self.ctrl_base + i, 0)
            # включить только выбранный
            self.reader.write_reg(self.ctrl_base + (fast_ch - 1), 1)
        finally:
            self.reader.write_reg(self.unlock_reg, 0)
        self._muted = True

    def _restore_channels(self):
        """Вернуть все каналы в 1, если мы их отключали."""
        if not self._muted or not (self.reader and self.is_connected):
            self._muted = False
            return
        try:
            self.reader.write_reg(self.unlock_reg, 1)
            for i in range(8):
                self.reader.write_reg(self.ctrl_base + i, 1)
        finally:
            self.reader.write_reg(self.unlock_reg, 0)
        self._muted = False

    # ---------- Опрос ----------
    def on_start(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return
        if not self.is_connected or not self.reader:
            messagebox.showerror("Ошибка", "Сначала подключитесь к устройству.")
            return

        try:
            fast_ch = max(1, min(8, int(self.fast_channel_var.get())))
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный номер канала.")
            return

        # Очистка и анти-ложный старт
        self.series_all.clear()
        self.trend_buffer.clear()
        self.value_label.configure(text="—")
        self._warmup_left = self._warmup_to_skip
        self._last_a_int = None

        # отключаем остальные каналы (и включаем выбранный)
        self._mute_others_and_enable_selected(fast_ch)

        self.stop_event.clear()
        self.worker_thread = threading.Thread(target=self._worker_loop, args=(fast_ch,), daemon=True)
        self.worker_thread.start()

        self.start_btn.configure(state=tk.DISABLED)
        self.stop_btn.configure(state=tk.NORMAL)
        self.status_var.set("Опрос запущен (10 мс)")
        self._timer_start()

    def _worker_loop(self, fast_channel):
        """1 канал, шаг 10 мс. Везде работаем в А (целые)."""
        address = fast_channel - 1
        interval_s = self.sample_interval_ms / 1000.0
        ui_update_s = 0.08
        next_ui_push = time.monotonic()

        try:
            while not self.stop_event.is_set():
                t0 = time.monotonic()
                raw = self.reader.read_ch(address, 1)
                ts = datetime.datetime.now()

                a_int = None
                if raw is not None and len(raw) >= 1:
                    a_int = self._code_to_A_int(raw[0])

                # пропускаем warmup + мусор
                if a_int is not None:
                    if self._warmup_left > 0:
                        self._warmup_left -= 1
                    else:
                        self.series_all.append((ts, a_int))
                        if time.monotonic() >= next_ui_push:
                            self.data_queue.put(("values", (ts, a_int)))
                            next_ui_push = time.monotonic() + ui_update_s
                        else:
                            # только в буфер тренда
                            line = [None]*8
                            line[address] = a_int
                            self.trend_buffer.append((ts, line))
                            if len(self.trend_buffer) > self.trend_buffer_max:
                                self.trend_buffer = self.trend_buffer[-self.trend_buffer_max:]

                dt = max(0.0, interval_s - (time.monotonic() - t0))
                if dt > 0:
                    time.sleep(dt)

        except Exception:
            pass
        finally:
            # При остановке не трогаем подключение, но возвращаем каналы
            self._restore_channels()
            self.data_queue.put(("status", "Подключено" if self.is_connected else "Отключено"))

    def on_stop(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.stop_event.set()
            self.worker_thread.join(timeout=2.0)
        self.start_btn.configure(state=tk.NORMAL if self.is_connected else tk.DISABLED)
        self.stop_btn.configure(state=tk.DISABLED)
        if self.is_connected:
            self.status_var.set("Опрос остановлен")
        self._timer_stop()

    def on_close(self):
        try:
            self.on_disconnect()
        finally:
            self.destroy()

    def on_clear_data(self):
        self.series_all.clear()
        self.trend_buffer.clear()
        self._last_a_int = None
        self.value_label.configure(text="—")
        self._redraw_trend()

    # ---------- Excel (Data + Changes + Trends по Changes) ----------
    def on_save_excel(self):
        if Workbook is None:
            messagebox.showerror("Не установлен openpyxl", "Для экспорта в Excel установите пакет:\n\npip install openpyxl")
            return
        if not self.series_all:
            messagebox.showwarning("Нет данных", "Пока нет накопленных данных для сохранения.")
            return

        fname = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            title="Сохранить как"
        )
        if not fname:
            return

        try:
            fast_idx = max(1, min(8, int(self.fast_channel_var.get())))

            # готовим массив (ts, A_int)
            rows = sorted(self.series_all, key=lambda x: x[0])
            if not rows:
                messagebox.showwarning("Нет данных", "Нет валидных точек для сохранения.")
                return

            # выборка «только изменения»
            changes = []
            last = None
            for ts, a in rows:
                if last is None or a != last:
                    changes.append((ts, a))
                    last = a
            if changes and changes[-1][0] != rows[-1][0]:
                changes.append(rows[-1])

            wb = Workbook()

            # --- Data: все точки (А) ---
            ws_data = wb.active
            ws_data.title = "Data"

            now = datetime.datetime.now()
            ws_data["A1"]  = "Дата/время сохранения:"; ws_data["B1"]  = now
            ws_data["A2"]  = "Порт:";                   ws_data["B2"]  = self.port_var.get()
            ws_data["A3"]  = "Скорость:";               ws_data["B3"]  = int(self.baud_var.get())
            ws_data["A4"]  = "Unit ID:";                ws_data["B4"]  = int(self.unit_var.get())
            ws_data["A5"]  = "Таймаут (с):";            ws_data["B5"]  = float(self.timeout_var.get())
            ws_data["A6"]  = "Канал:";                  ws_data["B6"]  = fast_idx
            ws_data["A7"]  = "Интервал опроса (мс):";   ws_data["B7"]  = self.sample_interval_ms
            ws_data["A8"]  = "Окно (с):";               ws_data["B8"]  = self.window_seconds
            ws_data["A9"]  = "Единицы:";                ws_data["B9"]  = "Амперы (целые)"

            header_row = 11
            ws_data.cell(row=header_row, column=1, value="Время")
            ws_data.cell(row=header_row, column=2, value=f"Канал{fast_idx} (A)")
            ws_data.cell(row=header_row, column=3, value="t, s")

            row_start = header_row + 1
            t0 = rows[0][0]
            for i, (ts, a) in enumerate(rows, start=row_start):
                c_time = ws_data.cell(row=i, column=1, value=ts)
                c_time.number_format = "yyyy-mm-dd hh:mm:ss.000"
                ws_data.cell(row=i, column=2, value=int(a))
                dt_s = (ts - t0).total_seconds()
                ws_data.cell(row=i, column=3, value=float(dt_s))

            # --- Changes: только изменения (А) ---
            ws_changes = wb.create_sheet("Changes")
            ws_changes["A1"] = "Время"
            ws_changes["B1"] = f"Канал{fast_idx} (A)"
            ws_changes["C1"] = "t, s"

            if changes:
                ch_start = 2
                for i, (ts, a) in enumerate(changes, start=ch_start):
                    c_time = ws_changes.cell(row=i, column=1, value=ts)
                    c_time.number_format = "yyyy-mm-dd hh:mm:ss.000"
                    ws_changes.cell(row=i, column=2, value=int(a))
                    ws_changes.cell(row=i, column=3, value=float((ts - t0).total_seconds()))
                ch_min_row = ch_start
                ch_max_row = ch_start + len(changes) - 1
            else:
                ch_min_row = ch_max_row = 2  # пусто — график не построится

            # --- Trends: график по листу Changes ---
            ws_chart = wb.create_sheet("Trends")
            if changes:
                chart = ScatterChart()
                chart.title = f"Channel {fast_idx} — only changes (A)"
                chart.style = 10
                chart.legend = None
                chart.x_axis.title = "Time, s"; chart.x_axis.number_format = "ss.000"
                chart.y_axis.title = "Current, A"

                # автошкала с паддингом
                y_vals = [a for _, a in changes]
                if y_vals:
                    y_min, y_max = min(y_vals), max(y_vals)
                    if y_min == y_max:
                        y_min -= 1; y_max += 1
                    pad = max(1, int(round((y_max - y_min) * 0.05)))
                    chart.y_axis.scaling.min = float(y_min - pad)
                    chart.y_axis.scaling.max = float(y_max + pad)

                x_ref = Reference(ws_changes, min_col=3, min_row=ch_min_row, max_row=ch_max_row)  # t,s
                y_ref = Reference(ws_changes, min_col=2, min_row=ch_min_row, max_row=ch_max_row)  # A
                s = Series(y_ref, xvalues=x_ref, title=None)
                try:
                    s.marker = None; s.smooth = True
                except Exception:
                    pass
                chart.series.append(s)
                chart.width = 28; chart.height = 12
                ws_chart.add_chart(chart, "A1")

            # --- Canvas → PNG → Excel (опционально)
            if PIL_AVAILABLE:
                try:
                    x = self.trend_canvas.winfo_rootx()
                    y = self.trend_canvas.winfo_rooty()
                    w = self.trend_canvas.winfo_width()
                    h = self.trend_canvas.winfo_height()
                    bbox = (x, y, x + w, y + h)
                    img = ImageGrab.grab(bbox)
                    img_path = fname.replace(".xlsx", "_trend.png")
                    img.save(img_path)
                    ws_img = wb.create_sheet("Trend Image")
                    ws_img.add_image(ExcelImage(img_path), "A1")
                except Exception:
                    pass

            wb.save(fname)
            messagebox.showinfo("Готово", f"Сохранено в файл:\n{fname}")

        except Exception as e:
            messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить Excel:\n{e}")


# -------------------
# Точка входа
# -------------------
if __name__ == "__main__":
    app = ModbusGUI()
    app.mainloop()
